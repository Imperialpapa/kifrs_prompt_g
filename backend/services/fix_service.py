"""
Fix Service - AI 스마트 수정 및 엑셀 변환
==========================================
오류에 대한 수정값을 제안하고, 엑셀 파일에 반영하는 서비스
"""

import io
import pandas as pd
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from datetime import datetime

from models import FixSuggestion, FixRequest, ValidationError
from database.validation_repository import ValidationRepository
from database.rule_repository import RuleRepository
from ai_layer import AIRuleInterpreter

class FixService:
    def __init__(self):
        self.validation_repo = ValidationRepository()
        self.rule_repo = RuleRepository()
        self.ai_interpreter = AIRuleInterpreter()

    async def suggest_fixes(self, session_id: str, error_ids: List[str] = None, provider: str = None) -> List[FixSuggestion]:
        """
        오류에 대한 AI 수정 제안 생성 (과거 이력 학습 적용)
        """
        print(f"[FixService] Generating suggestions for session: {session_id} using {provider}")
        
        try:
            # 1. 현재 오류 데이터 조회
            query = self.validation_repo.client.table('validation_errors') \
                .select('*') \
                .eq('session_id', session_id)
            
            if error_ids:
                query = query.in_('id', error_ids)
            
            error_result = query.limit(100).execute()
            errors = error_result.data
            
            if not errors:
                return []
                
            # 2. 과거 수정 이력 조회 (Learning / RAG)
            # 최근 50개의 성공적인 수정 사례를 가져와 AI에게 '학습' 시킴
            past_corrections = []
            try:
                history_result = self.validation_repo.client.table('user_corrections') \
                    .select('column_name, old_value, new_value, correction_reason') \
                    .order('created_at', desc=True) \
                    .limit(50).execute()
                past_corrections = history_result.data
                print(f"[FixService] Loaded {len(past_corrections)} past correction examples for learning")
            except Exception as e:
                print(f"[FixService] History lookup failed (non-critical): {e}")

            # 3. AI/로컬 엔진 호출
            formatted_errors = [{
                "id": err['id'],
                "sheet": err['sheet_name'],
                "row": err['row_number'],
                "column": err['column_name'],
                "actual_value": err['actual_value'],
                "message": err['error_message']
            } for err in errors]

            suggestions = await self.ai_interpreter.suggest_corrections(
                formatted_errors, 
                past_corrections,
                provider=provider
            )
            
            return suggestions

        except Exception as e:
            print(f"[FixService] Critical error in suggest_fixes: {e}")
            return []

    def apply_fixes_to_excel(
        self, 
        original_file_content: bytes, 
        fixes: List[FixRequest]
    ) -> bytes:
        """
        원본 엑셀 파일에 수정 사항을 반영
        
        Args:
            original_file_content: 업로드된 원본 엑셀 파일 (Binary)
            fixes: 적용할 수정 목록
            
        Returns:
            bytes: 수정된 엑셀 파일
        """
        print(f"[FixService] Applying {len(fixes)} fixes to Excel...")
        
        # openpyxl로 로드 (data_only=False로 수식 유지, but 값 수정 시 주의)
        wb = load_workbook(io.BytesIO(original_file_content))
        
        # 빠른 조회를 위해 (sheet, row, col) -> fix 맵핑 생성
        fix_map = {}
        for fix in fixes:
            # 엑셀의 컬럼명(A, B, C...)을 인덱스로 변환하는 로직 필요할 수 있음
            # 현재 모델은 column이 '이름', '생년월일' 등 헤더명일 수 있음.
            # 따라서 Row 값은 엑셀의 물리적 행 번호(1-based)여야 정확함.
            key = (fix.sheet_name, fix.row, fix.column)
            fix_map[key] = fix.fixed_value

        # 시트별로 순회하며 수정
        # 주의: column이 '헤더명'인 경우, 해당 헤더가 몇 번째 열인지 찾아야 함.
        # 성능을 위해 미리 헤더 맵을 만드는 것이 좋음.
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 헤더 매핑 (1행이 헤더로 가정)
            # TODO: 실제 헤더 위치(row)는 메타데이터에서 가져와야 정확함. 
            # Phase 4의 parser 로직 참고. 여기서는 단순화하여 1~3행 스캔.
            
            header_map = {} # '성명' -> 3 (C열)
            header_row_idx = None
            
            # 헤더 찾기 (간이 로직)
            for r in range(1, 6):
                row_values = [c.value for c in ws[r]]
                if any(v for v in row_values if isinstance(v, str)):
                    header_row_idx = r
                    for c_idx, cell in enumerate(ws[r], start=1):
                        if cell.value:
                            # 정규화된 이름으로 매핑 (공백 제거 등)
                            clean_header = str(cell.value).strip()
                            header_map[clean_header] = c_idx
                            # 원본 이름으로도 매핑
                            header_map[str(cell.value)] = c_idx
                    break
            
            if not header_row_idx:
                print(f"[FixService] Warning: Could not find header for sheet '{sheet_name}'")
                continue

            # 수정 적용
            applied_count = 0
            for fix in fixes:
                if fix.sheet_name != sheet_name:
                    continue
                
                # 타겟 컬럼 인덱스 찾기
                col_idx = header_map.get(fix.column)
                if not col_idx:
                    # 혹시 fix.column이 이미 'A', 'B' 형태라면? (Phase 4 parser는 column_letter를 줌)
                    # 여기서는 column이 헤더명이라고 가정.
                    continue
                
                # 셀 업데이트
                try:
                    cell = ws.cell(row=fix.row, column=col_idx)
                    # 스타일 유지하며 값만 변경
                    original_val_in_cell = cell.value
                    cell.value = fix.fixed_value
                    applied_count += 1
                    # print(f"  - Updated {sheet_name}!{fix.column}{fix.row}: {original_val_in_cell} -> {fix.fixed_value}")
                except Exception as e:
                    print(f"  - Failed to update cell: {e}")

            print(f"[FixService] Sheet '{sheet_name}': {applied_count} changes applied")

        # 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def apply_bulk_fixes_to_excel(
        self,
        original_file_content: bytes,
        cells_to_fix: List[Dict[str, Any]],
        filename: str = ""
    ) -> bytes:
        """
        컬럼별 일괄 수정 적용 및 변경내역 시트 추가

        Args:
            original_file_content: 원본 엑셀 파일 바이트
            cells_to_fix: 수정할 셀 목록 [{sheet, row, column, currentValue, fixType}, ...]
            filename: 원본 파일명 (확장자 판단용)

        Returns:
            bytes: 수정된 엑셀 파일
        """
        from openpyxl.styles import PatternFill
        from openpyxl import Workbook

        print(f"[FixService] Bulk fixing {len(cells_to_fix)} cells... (file: {filename})")

        # 파일 확장자 확인
        is_xls = filename.lower().endswith('.xls') and not filename.lower().endswith('.xlsx')

        wb = None
        
        if is_xls:
            # .xls 파일 처리: 서식 유지가 어려우므로 변환 후 값만 유지
            print("[FixService] Detected .xls format. Converting to .xlsx via pandas (Formatting may be lost)...")
            try:
                # pandas로 xls 읽기
                xls_data = pd.read_excel(io.BytesIO(original_file_content), sheet_name=None, engine='xlrd')

                # 새 workbook 생성 (서식 초기화됨)
                wb = Workbook()
                wb.remove(wb.active)  # 기본 시트 제거

                for sheet_name, df in xls_data.items():
                    ws = wb.create_sheet(title=sheet_name)
                    # 헤더 쓰기
                    for col_idx, col_name in enumerate(df.columns, start=1):
                        ws.cell(row=1, column=col_idx, value=col_name)
                    # 데이터 쓰기
                    for row_idx, row in enumerate(df.values, start=2):
                        for col_idx, value in enumerate(row, start=1):
                            # NaN 처리
                            if pd.isna(value):
                                ws.cell(row=row_idx, column=col_idx, value=None)
                            else:
                                ws.cell(row=row_idx, column=col_idx, value=value)

                print(f"[FixService] Converted {len(xls_data)} sheets from xls")
            except Exception as e:
                print(f"[FixService] xls conversion failed: {e}")
                raise ValueError(f"Failed to process .xls file: {str(e)}")
        else:
            # .xlsx 파일 처리: 원본 파일을 그대로 로드하여 수정 (서식 유지)
            print("[FixService] Detected .xlsx format. Loading original file to preserve formatting...")
            try:
                # BytesIO를 통해 메모리상의 원본 파일을 직접 로드
                # openpyxl은 이 시점에서 파일 구조와 스타일을 메모리에 적재함
                wb = load_workbook(io.BytesIO(original_file_content))
            except Exception as e:
                print(f"[FixService] Failed to load .xlsx file: {e}")
                raise ValueError(f"Failed to load .xlsx file: {str(e)}")

        # 수정 결과 추적
        column_stats = {}  # column -> {sheets: set, count, success, fail}

        # 스타일 정의 (수정된 셀 표시용)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

        # 시트별 헤더 매핑 캐시
        header_maps = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_map = {}

            # 헤더 찾기 (1~5행 중 첫 번째 텍스트 행)
            for r in range(1, 6):
                row_values = [c.value for c in ws[r]]
                if any(v for v in row_values if isinstance(v, str)):
                    for c_idx, cell in enumerate(ws[r], start=1):
                        if cell.value:
                            clean_header = str(cell.value).strip()
                            header_map[clean_header] = c_idx
                    break

            header_maps[sheet_name] = header_map

        # 각 셀 수정 적용
        for cell_info in cells_to_fix:
            sheet = cell_info.get('sheet', '기본')
            row = cell_info.get('row')
            column = cell_info.get('column')
            # currentValue = cell_info.get('currentValue') # Unused
            fix_type = cell_info.get('fixType', 'unknown')

            if sheet not in wb.sheetnames:
                continue

            ws = wb[sheet]
            header_map = header_maps.get(sheet, {})
            col_idx = header_map.get(column)

            if not col_idx:
                continue

            # 통계 초기화
            if column not in column_stats:
                column_stats[column] = {
                    'sheets': set(),
                    'fix_type': fix_type,
                    'success': 0,
                    'fail': 0,
                    'samples': []
                }

            column_stats[column]['sheets'].add(sheet)

            # 값 변환 및 셀 업데이트
            try:
                cell = ws.cell(row=row, column=col_idx)
                original_value = cell.value
                fixed_value = self._convert_value(original_value, fix_type)

                if fixed_value is not None and fixed_value != original_value:
                    # 값 수정
                    cell.value = fixed_value
                    
                    # 수정 표시 (배경색 변경)
                    # 주의: 원본 셀의 스타일(테두리, 폰트 등)은 유지되지만 배경색은 덮어씌워짐
                    cell.fill = yellow_fill  
                    
                    column_stats[column]['success'] += 1

                    # 샘플 저장 (최대 3개)
                    if len(column_stats[column]['samples']) < 3:
                        column_stats[column]['samples'].append({
                            'before': str(original_value),
                            'after': str(fixed_value)
                        })
                else:
                    # 수정 실패 표시
                    cell.fill = red_fill
                    column_stats[column]['fail'] += 1

            except Exception as e:
                print(f"[FixService] Error fixing cell {sheet}:{column}:{row} - {e}")
                column_stats[column]['fail'] += 1

        # --- [LEARNING START] Save corrections to DB ---
        try:
            corrections_to_save = []
            # We don't have session_id here easily, so we might use a placeholder or omit.
            # But the table structure expects it. Let's assume we can skip it or generate one?
            # Actually, `apply_bulk_fixes_to_excel` doesn't receive session_id.
            # For now, let's store general patterns without specific session link if nullable,
            # or skip session_id if schema allows.
            
            # Since we iterate by column statistics, let's save representative examples
            for column, stats in column_stats.items():
                if stats['success'] > 0 and stats['samples']:
                    # Save a few examples for this column
                    for sample in stats['samples']:
                        corrections_to_save.append({
                            "session_id": "00000000-0000-0000-0000-000000000000", # Placeholder
                            "sheet_name": list(stats['sheets'])[0], # Just one sheet as example
                            "column_name": column,
                            "old_value": sample['before'][:255], # Truncate if too long
                            "new_value": sample['after'][:255],
                            "correction_action": "bulk_fix",
                            "correction_reason": f"Bulk fix: {stats['fix_type']}",
                            "corrected_by": "user",
                            "created_at": datetime.now().isoformat()
                        })
            
            if corrections_to_save:
                # Use rule_repo's client or validation_repo's client
                self.validation_repo.client.table('user_corrections').insert(corrections_to_save).execute()
                print(f"[FixService] Learned {len(corrections_to_save)} correction patterns from bulk fix.")
                
        except Exception as e:
            print(f"[FixService] Failed to save learning data (non-critical): {e}")
        # --- [LEARNING END] ---

        # _변경내역 시트 생성 (항상 맨 뒤에 추가)
        change_sheet_name = "_변경내역"
        if change_sheet_name in wb.sheetnames:
            # 기존 변경내역 시트가 있다면 삭제하고 재생성 (누적 방지)
            del wb[change_sheet_name]

        ws_log = wb.create_sheet(change_sheet_name)

        # 헤더 작성
        headers = ["컬럼명", "수정 규칙", "원본 예시", "수정 예시", "적용 시트", "수정 건수", "상태"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_log.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            from openpyxl.styles import Font
            cell.font = Font(color="FFFFFF", bold=True)

        # 데이터 작성
        row_idx = 2
        for column, stats in column_stats.items():
            fix_desc = self._get_fix_description(stats['fix_type'])

            # 원본/수정 예시
            before_examples = ", ".join([s['before'] for s in stats['samples'][:3]]) if stats['samples'] else "-"
            after_examples = ", ".join([s['after'] for s in stats['samples'][:3]]) if stats['samples'] else "-"

            # 적용 시트
            sheets_str = ", ".join(sorted(stats['sheets']))

            # 상태
            # total = stats['success'] + stats['fail'] # Unused
            if stats['fail'] == 0:
                status = "✅ 수정완료"
            elif stats['success'] == 0:
                status = "❌ 수정실패"
            else:
                status = f"⚠️ 일부실패 ({stats['fail']}건)"

            ws_log.cell(row=row_idx, column=1, value=column)
            ws_log.cell(row=row_idx, column=2, value=fix_desc)
            ws_log.cell(row=row_idx, column=3, value=before_examples)
            ws_log.cell(row=row_idx, column=4, value=after_examples)
            ws_log.cell(row=row_idx, column=5, value=sheets_str)
            ws_log.cell(row=row_idx, column=6, value=f"{stats['success']}건")
            ws_log.cell(row=row_idx, column=7, value=status)

            row_idx += 1

        # 컬럼 너비 자동 조정
        ws_log.column_dimensions['A'].width = 15
        ws_log.column_dimensions['B'].width = 30
        ws_log.column_dimensions['C'].width = 25
        ws_log.column_dimensions['D'].width = 25
        ws_log.column_dimensions['E'].width = 25
        ws_log.column_dimensions['F'].width = 12
        ws_log.column_dimensions['G'].width = 15

        # 저장 (BytesIO로 출력)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        print(f"[FixService] Bulk fix complete. Modified {sum(s['success'] for s in column_stats.values())} cells.")

        return output.getvalue()

    def _convert_value(self, value: Any, fix_type: str) -> Any:
        """
        값 변환 로직
        """
        if value is None:
            return None

        # datetime 객체인 경우 (최우선 처리)
        if isinstance(value, datetime):
            return value.strftime('%Y%m%d')

        str_value = str(value).strip()

        if fix_type == 'date_format':
            # 1. 시리얼 번호(숫자)인 경우 처리 (예: 45292)
            try:
                num_val = float(value)
                # 시리얼 날짜 범위 (1 ~ 100,000 -> 1900년 ~ 2173년)
                if 1 <= num_val <= 100000: 
                    dt = pd.to_datetime(num_val, unit='D', origin='1899-12-30')
                    return dt.strftime('%Y%m%d')
                
                # 이미 YYYYMMDD 숫자인 경우 (예: 19881115.0) -> 문자열로 변환
                if 19000101 <= num_val <= 21001231:
                    return str(int(num_val))
            except (ValueError, TypeError):
                pass

            # 2. 구분자가 있는 문자열 처리 (YYYY-MM-DD, YYYY.MM.DD 등)
            import re
            # Loose matching: allow spaces, any separator, ignore trailing chars
            match = re.match(r'^(\d{4})\s*[-/.]\s*(\d{1,2})\s*[-/.]\s*(\d{1,2})', str_value)
            if match:
                year = match.group(1)
                month = match.group(2).zfill(2) # Pad with 0
                day = match.group(3).zfill(2)   # Pad with 0
                return f"{year}{month}{day}"

            return value

        elif fix_type == 'gender_code':
            # 남/여/남자/여자/M/F → 1/2
            gender_map = {
                '남': '1', '남자': '1', 'm': '1', 'male': '1', 'M': '1',
                '여': '2', '여자': '2', 'f': '2', 'female': '2', 'F': '2'
            }
            return gender_map.get(str_value, value)

        elif fix_type == 'number_format':
            # 콤마, 공백 제거
            import re
            cleaned = re.sub(r'[,\s]', '', str_value)
            if cleaned.isdigit():
                return int(cleaned)
            try:
                return float(cleaned)
            except ValueError:
                return value

        elif fix_type == 'trim':
            return str_value

        return value

    def _get_fix_description(self, fix_type: str) -> str:
        """
        수정 유형 설명
        """
        descriptions = {
            'date_format': '날짜 형식 변환 (YYYY-MM-DD → YYYYMMDD)',
            'gender_code': '성별 코드 변환 (남→1, 여→2)',
            'number_format': '숫자 형식 변환 (콤마/공백 제거)',
            'trim': '앞뒤 공백 제거',
            'duplicate': '중복 오류 (자동수정 불가)',
            'required': '필수값 누락 (자동수정 불가)',
            'unknown': '알 수 없는 수정 유형'
        }
        return descriptions.get(fix_type, fix_type)
