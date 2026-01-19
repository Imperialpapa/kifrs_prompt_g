"""
Rule Service - Business Logic Layer
====================================
규칙 파일 업로드, 다운로드, 조회 등의 비즈니스 로직 처리
"""

import io
from typing import List, Dict, Any, Optional
from uuid import UUID
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database.rule_repository import RuleRepository
from utils.excel_parser import parse_rules_from_excel, normalize_sheet_name, get_canonical_name
from models import RuleFileUpload, RuleFileResponse, RuleCreate
from services.ai_cache_service import AICacheService


class RuleService:
    """
    Service layer for rule management

    Provides high-level business logic for:
    - Uploading rule files to database
    - Exporting rules from database to Excel
    - Retrieving rule file details and statistics
    """

    def __init__(self):
        """Initialize service with repository"""
        self.repository = RuleRepository()
        self.ai_cache_service = AICacheService()

    async def upload_rule_file(
        self,
        excel_content: bytes,
        metadata: RuleFileUpload
    ) -> RuleFileResponse:
        """
        Upload and parse rule file to database

        Process:
        1. Parse Excel B file using excel_parser
        2. Create rule_file record in database
        3. Batch insert all rules
        4. Return metadata response

        Args:
            excel_content: Raw Excel file bytes
            metadata: File metadata (file_name, version, etc.)

        Returns:
            RuleFileResponse: Created rule file metadata

        Raises:
            Exception: If parsing or database operations fail
        """
        print(f"[RuleService] Starting upload for file: {metadata.file_name}")

        try:
            # Step 1: Parse Excel file
            print("[RuleService] Parsing Excel file...")
            natural_language_rules, sheet_row_counts, total_raw_rows, reported_max_row = parse_rules_from_excel(excel_content)

            print(f"[RuleService] Parsed {len(natural_language_rules)} rules from {len(sheet_row_counts)} sheets")

            # Step 2: Prepare rules for batch insert FIRST (before creating file record)
            print("[RuleService] Preparing rules for batch insert...")
            rules_to_insert = []

            for rule in natural_language_rules:
                rule_record = {
                    "sheet_name": rule["display_sheet_name"],  # Required by DB schema
                    "canonical_sheet_name": rule["sheet"],
                    "display_sheet_name": rule["display_sheet_name"],
                    "row_number": rule["row"],
                    "column_letter": rule["column_letter"],
                    "field_name": rule["field"],
                    "rule_text": rule["rule_text"],
                    "condition": rule["condition"],
                    "note": rule["note"],
                    "is_active": True
                }
                rules_to_insert.append(rule_record)

            print(f"[RuleService] Prepared {len(rules_to_insert)} rules for insertion")

            # Step 3: Check for duplicate file (same name + version)
            requested_version = metadata.file_version or "1.0"
            print(f"[RuleService] Checking for existing file: {metadata.file_name} v{requested_version}")

            existing_files = await self.repository.list_rule_files(status='active', limit=100)
            duplicate_found = False
            max_version_number = 0.0

            for existing in existing_files:
                if existing['file_name'] == metadata.file_name:
                    # Parse version as float
                    try:
                        existing_version = float(existing['file_version'])
                        max_version_number = max(max_version_number, existing_version)

                        if existing['file_version'] == requested_version:
                            duplicate_found = True
                            print(f"[RuleService] Found duplicate: {existing['id']} (uploaded {existing['uploaded_at']})")
                    except ValueError:
                        pass

            # Auto-increment version if duplicate found
            final_version = requested_version
            if duplicate_found:
                new_version_number = max_version_number + 0.1
                final_version = f"{new_version_number:.1f}"
                print(f"[RuleService] Auto-incrementing version: {requested_version} → {final_version}")

            # Step 4: Create rule_file record
            file_data = {
                "file_name": metadata.file_name,
                "file_version": final_version,
                "uploaded_by": metadata.uploaded_by or "system",
                "total_rules_count": len(natural_language_rules),
                "sheet_count": len(sheet_row_counts),
                "notes": metadata.notes,
                "status": "active"
            }

            print("[RuleService] Creating rule_file record...")
            file_id = await self.repository.create_rule_file(file_data)
            print(f"[RuleService] Created rule_file with ID: {file_id}")

            # Step 5: Add file_id to all rules and batch insert
            print(f"[RuleService] Adding file_id to rules and inserting...")
            for rule_record in rules_to_insert:
                rule_record["rule_file_id"] = file_id

            try:
                inserted_count = await self.repository.create_rules_batch(rules_to_insert)
                print(f"[RuleService] Successfully inserted {inserted_count} rules")

                # Verify insertion
                if inserted_count == 0:
                    raise Exception("Batch insert returned 0 rows - rules may not have been saved")

                if inserted_count != len(rules_to_insert):
                    print(f"[RuleService] WARNING: Expected {len(rules_to_insert)} rules, but inserted {inserted_count}")

            except Exception as insert_error:
                print(f"[RuleService] CRITICAL: Failed to insert rules: {str(insert_error)}")
                print(f"[RuleService] Attempting to delete orphaned file record: {file_id}")

                # Try to clean up the file record
                try:
                    await self.repository.archive_rule_file(file_id)
                    print(f"[RuleService] Archived orphaned file record")
                except Exception as cleanup_error:
                    print(f"[RuleService] Failed to clean up file record: {str(cleanup_error)}")

                raise Exception(f"Failed to insert rules to database: {str(insert_error)}")

            # Step 6: Save original file for future re-interpretation
            print("[RuleService] Saving original file content...")
            save_success = await self.repository.save_original_file(UUID(file_id), excel_content)
            if save_success:
                print("[RuleService] Original file saved successfully")
            else:
                print("[RuleService] Warning: Failed to save original file (non-critical)")

            # Step 7: Retrieve and return file metadata
            file_record = await self.repository.get_rule_file(UUID(file_id))

            response = RuleFileResponse(
                id=file_record['id'],
                file_name=file_record['file_name'],
                file_version=file_record.get('file_version'),
                uploaded_by=file_record.get('uploaded_by'),
                uploaded_at=file_record['uploaded_at'],
                sheet_count=file_record['sheet_count'],
                total_rules_count=file_record['total_rules_count'],
                status=file_record['status']
            )

            print(f"[RuleService] Upload completed successfully")

            # Step 8: Start rule interpretation (use local parser by default to avoid AI errors)
            print(f"[RuleService] Starting rule interpretation (local parser)...")
            try:
                # Use local parser by default to avoid AI interpretation errors
                ai_result = await self.ai_cache_service.interpret_and_cache_rules(
                    file_id,
                    force_local=True  # 로컬 파서 사용 (AI 오류 방지)
                )
                print(f"[RuleService] Interpretation completed: {ai_result['interpreted_rules']}/{ai_result['total_rules']} rules (engine: {ai_result.get('engine', 'local')})")
            except Exception as ai_error:
                print(f"[RuleService] Interpretation failed (non-critical): {str(ai_error)}")
                # Don't fail the upload if interpretation fails

            return response

        except Exception as e:
            print(f"[RuleService] Error during upload: {str(e)}")
            raise Exception(f"Failed to upload rule file: {str(e)}")

    async def create_single_rule(self, rule_data: RuleCreate) -> str:
        """
        Create a single rule manually

        Args:
            rule_data: Rule creation data

        Returns:
            str: ID of the created rule
        """
        print(f"[RuleService] Creating single rule for file: {rule_data.rule_file_id}")
        try:
            # Prepare rule record
            rule_record = {
                "rule_file_id": rule_data.rule_file_id,
                "sheet_name": rule_data.sheet_name,
                "canonical_sheet_name": get_canonical_name(rule_data.sheet_name),
                "display_sheet_name": rule_data.sheet_name,
                "row_number": rule_data.row_number,
                "column_letter": "Manual",  # Placeholder
                "field_name": rule_data.column_name,
                "rule_text": rule_data.rule_text,
                "condition": rule_data.condition,
                "is_active": True,
                "is_common": rule_data.is_common,
                # AI fields (manual input or default)
                "ai_rule_type": rule_data.ai_rule_type,
                "ai_parameters": rule_data.ai_parameters,
                "ai_confidence_score": 1.0,  # User manually created -> 100% confidence
                "ai_interpretation_summary": "사용자 직접 입력 규칙",
                "ai_model_version": "manual"
            }

            # Use the new repository method
            created_rule = await self.repository.create_single_rule(rule_record)
            
            if created_rule and created_rule.get('id'):
                # Update total_rules_count in file record
                await self.repository.increment_rule_count(UUID(rule_data.rule_file_id))
                return created_rule['id'] # Return the new ID
            else:
                raise Exception("Failed to insert rule record or get ID back")

        except Exception as e:
            print(f"[RuleService] Error creating single rule: {str(e)}")
            raise Exception(f"Failed to create single rule: {str(e)}")

    async def get_rule_file_details(self, file_id: str) -> Optional[Dict[str, Any]]:
        """
        Get detailed information about a rule file

        Args:
            file_id: UUID string of the rule file

        Returns:
            Dict with file metadata and statistics, or None if not found
        """
        print(f"[RuleService] Fetching details for file: {file_id}")

        try:
            # Get file metadata
            file_record = await self.repository.get_rule_file(UUID(file_id))

            if not file_record:
                print(f"[RuleService] File not found: {file_id}")
                return None

            # Get statistics
            stats = await self.repository.get_file_statistics(UUID(file_id))

            # Get rules grouped by sheet
            all_rules = await self.repository.get_rules_by_file(UUID(file_id), active_only=True)

            # Group rules by sheet
            from collections import defaultdict
            rules_by_sheet = defaultdict(list)
            for rule in all_rules:
                sheet_name = rule.get('display_sheet_name', 'Unknown')
                rules_by_sheet[sheet_name].append({
                    "id": str(rule.get('id')),
                    "field_name": rule.get('field_name'),
                    "rule_text": rule.get('rule_text'),
                    "is_common": rule.get('is_common', False),
                    "has_ai_interpretation": bool(rule.get('ai_rule_id'))
                })

            # Build response
            response = {
                "id": file_record['id'],
                "file_name": file_record['file_name'],
                "file_version": file_record.get('file_version'),
                "uploaded_by": file_record.get('uploaded_by'),
                "uploaded_at": file_record['uploaded_at'],
                "updated_at": file_record.get('updated_at'),
                "sheet_count": file_record['sheet_count'],
                "total_rules_count": file_record['total_rules_count'],
                "status": file_record['status'],
                "notes": file_record.get('notes'),
                "statistics": stats,
                "sheets": [
                    {
                        "sheet_name": sheet_name,
                        "rule_count": len(rules),
                        "sample_rules": rules[:5]  # Show first 5 rules as sample
                    }
                    for sheet_name, rules in rules_by_sheet.items()
                ]
            }

            print(f"[RuleService] Retrieved details for file: {file_id}")
            return response

        except Exception as e:
            print(f"[RuleService] Error fetching file details: {str(e)}")
            raise Exception(f"Failed to get rule file details: {str(e)}")

    async def get_rule_mappings(self, file_id: str) -> Optional[Dict[str, Any]]:
        """
        Get AI mapping details for all rules in a file

        Args:
            file_id: UUID string of the rule file

        Returns:
            Dict with mapping statistics and detailed rule mappings
        """
        print(f"[RuleService] Fetching rule mappings for file: {file_id}")

        try:
            # Get file metadata
            file_record = await self.repository.get_rule_file(UUID(file_id))

            if not file_record:
                print(f"[RuleService] File not found: {file_id}")
                return None

            # Get all rules
            all_rules = await self.repository.get_rules_by_file(UUID(file_id), active_only=True)

            # Categorize rules by mapping status
            mapped_rules = []
            unmapped_rules = []
            partial_rules = []  # Has AI interpretation but low confidence

            for rule in all_rules:
                rule_info = {
                    "id": str(rule.get('id')),
                    "sheet_name": rule.get('display_sheet_name', 'Unknown'),
                    "row_number": rule.get('row_number'),
                    "column_letter": rule.get('column_letter'),
                    "field_name": rule.get('field_name'),
                    "rule_text": rule.get('rule_text'),
                    "condition": rule.get('condition'),
                    "note": rule.get('note'),
                    "is_common": rule.get('is_common', False),
                    # AI interpretation fields
                    "ai_rule_id": rule.get('ai_rule_id'),
                    "ai_rule_type": rule.get('ai_rule_type'),
                    "ai_parameters": rule.get('ai_parameters'),
                    "ai_error_message": rule.get('ai_error_message'),
                    "ai_interpretation_summary": rule.get('ai_interpretation_summary'),
                    "ai_confidence_score": rule.get('ai_confidence_score'),
                    "ai_model_version": rule.get('ai_model_version'),
                    "is_active": rule.get('is_active', True)
                }

                # Determine mapping status
                if rule.get('ai_rule_id') and rule.get('ai_rule_type'):
                    confidence = rule.get('ai_confidence_score', 0)
                    if confidence and confidence >= 0.8:
                        rule_info["mapping_status"] = "mapped"
                        mapped_rules.append(rule_info)
                    elif confidence and confidence >= 0.5:
                        rule_info["mapping_status"] = "partial"
                        partial_rules.append(rule_info)
                    else:
                        rule_info["mapping_status"] = "mapped"
                        mapped_rules.append(rule_info)
                else:
                    rule_info["mapping_status"] = "unmapped"
                    unmapped_rules.append(rule_info)

            # Group by sheet for easier display
            from collections import defaultdict
            rules_by_sheet = defaultdict(list)
            for rule in mapped_rules + partial_rules + unmapped_rules:
                rules_by_sheet[rule['sheet_name']].append(rule)

            # Sort rules within each sheet by row number
            for sheet_name in rules_by_sheet:
                rules_by_sheet[sheet_name].sort(key=lambda x: x.get('row_number', 0))

            # Build response
            total_rules = len(all_rules)
            mapped_count = len(mapped_rules)
            partial_count = len(partial_rules)
            unmapped_count = len(unmapped_rules)

            response = {
                "file_id": file_id,
                "file_name": file_record['file_name'],
                "statistics": {
                    "total_rules": total_rules,
                    "mapped_count": mapped_count,
                    "partial_count": partial_count,
                    "unmapped_count": unmapped_count,
                    "mapping_rate": round((mapped_count + partial_count) / total_rules * 100, 1) if total_rules > 0 else 0,
                    "full_mapping_rate": round(mapped_count / total_rules * 100, 1) if total_rules > 0 else 0
                },
                "sheets": [
                    {
                        "sheet_name": sheet_name,
                        "rules": rules,
                        "mapped_count": sum(1 for r in rules if r['mapping_status'] == 'mapped'),
                        "partial_count": sum(1 for r in rules if r['mapping_status'] == 'partial'),
                        "unmapped_count": sum(1 for r in rules if r['mapping_status'] == 'unmapped')
                    }
                    for sheet_name, rules in sorted(rules_by_sheet.items())
                ],
                "available_rule_types": [
                    {"value": "required", "label": "필수 입력 (required)", "description": "빈 값 또는 NULL 검사"},
                    {"value": "format", "label": "형식 검증 (format)", "description": "정규식, 허용값 목록, 날짜 형식 등"},
                    {"value": "range", "label": "범위 검증 (range)", "description": "숫자/날짜 최소/최대값"},
                    {"value": "no_duplicates", "label": "중복 금지 (no_duplicates)", "description": "고유값 검사"},
                    {"value": "composite", "label": "복합 검증 (composite)", "description": "여러 검증 조건을 한번에 적용 (필수+형식 등)"},
                    {"value": "date_logic", "label": "날짜 논리 (date_logic)", "description": "날짜 간 비교 (입사일 > 생년월일)"},
                    {"value": "cross_field", "label": "교차 필드 (cross_field)", "description": "필드 간 상호 검증"},
                    {"value": "custom", "label": "사용자 정의 (custom)", "description": "복잡한 비즈니스 로직"}
                ]
            }

            print(f"[RuleService] Retrieved mappings: {mapped_count} mapped, {partial_count} partial, {unmapped_count} unmapped")
            return response

        except Exception as e:
            print(f"[RuleService] Error fetching rule mappings: {str(e)}")
            raise Exception(f"Failed to get rule mappings: {str(e)}")

    async def list_rule_files(
        self,
        status: str = 'active',
        limit: int = 50,
        offset: int = 0
    ) -> List[RuleFileResponse]:
        """
        List all rule files with pagination

        Args:
            status: Filter by status (default: 'active')
            limit: Maximum number of results
            offset: Pagination offset

        Returns:
            List of RuleFileResponse objects
        """
        print(f"[RuleService] Listing rule files (status={status}, limit={limit}, offset={offset})")

        try:
            files = await self.repository.list_rule_files(status, limit, offset)

            response_list = [
                RuleFileResponse(
                    id=file['id'],
                    file_name=file['file_name'],
                    file_version=file.get('file_version'),
                    uploaded_by=file.get('uploaded_by'),
                    uploaded_at=file['uploaded_at'],
                    sheet_count=file['sheet_count'],
                    total_rules_count=file['total_rules_count'],
                    status=file['status']
                )
                for file in files
            ]

            print(f"[RuleService] Found {len(response_list)} rule files")
            return response_list

        except Exception as e:
            print(f"[RuleService] Error listing files: {str(e)}")
            raise Exception(f"Failed to list rule files: {str(e)}")

    async def get_rule(self, rule_id: str) -> Optional[Dict[str, Any]]:
        """
        Get a single rule by ID

        Args:
            rule_id: UUID string of the rule

        Returns:
            Dict or None
        """
        try:
            rule = await self.repository.get_rule(UUID(rule_id))
            return rule
        except Exception as e:
            print(f"[RuleService] Error getting rule: {str(e)}")
            raise Exception(f"Failed to get rule: {str(e)}")

    async def update_rule(self, rule_id: str, updates: Dict[str, Any]) -> bool:
        """
        Update a rule

        Args:
            rule_id: UUID string of the rule
            updates: Dictionary of fields to update

        Returns:
            bool: True if successful
        """
        print(f"[RuleService] Updating rule: {rule_id}")
        try:
            # Clean updates: remove None values to avoid overwriting with None
            # (Unless we specifically want to set something to None)
            clean_updates = {k: v for k, v in updates.items() if v is not None}
            
            success = await self.repository.update_rule(UUID(rule_id), clean_updates)
            return success
        except Exception as e:
            print(f"[RuleService] Error updating rule: {str(e)}")
            raise Exception(f"Failed to update rule: {str(e)}")

    async def delete_rule(self, rule_id: str, permanent: bool = False) -> bool:
        """
        Delete or deactivate a rule

        Args:
            rule_id: UUID string of the rule
            permanent: If True, delete from DB. If False, just deactivate.

        Returns:
            bool: True if successful
        """
        print(f"[RuleService] Deleting rule: {rule_id} (permanent={permanent})")
        try:
            if permanent:
                return await self.repository.delete_rule(UUID(rule_id))
            else:
                return await self.repository.deactivate_rule(UUID(rule_id))
        except Exception as e:
            print(f"[RuleService] Error deleting rule: {str(e)}")
            raise Exception(f"Failed to delete rule: {str(e)}")

    async def archive_rule_file(self, file_id: str) -> bool:
        """
        Archive a rule file (soft delete)

        Args:
            file_id: UUID string of the rule file

        Returns:
            bool: True if successful
        """
        print(f"[RuleService] Archiving rule file: {file_id}")
        try:
            return await self.repository.archive_rule_file(UUID(file_id))
        except Exception as e:
            print(f"[RuleService] Error archiving rule file: {str(e)}")
            raise Exception(f"Failed to archive rule file: {str(e)}")

    async def export_rules_to_excel(self, file_id: str) -> bytes:
        """
        Export rules from database back to Excel format

        Args:
            file_id: UUID string of the rule file

        Returns:
            bytes: Excel file content

        Raises:
            Exception: If file not found or export fails
        """
        print(f"[RuleService] Exporting rules to Excel for file: {file_id}")

        try:
            # Get file metadata
            file_record = await self.repository.get_rule_file(UUID(file_id))

            if not file_record:
                raise Exception(f"Rule file not found: {file_id}")

            # Get all rules
            all_rules = await self.repository.get_rules_by_file(UUID(file_id), active_only=True)

            if not all_rules:
                raise Exception(f"No rules found for file: {file_id}")

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "규칙 목록"

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Write headers (Row 1: Main header, Row 2: Sub header - matches parser min_row=3)
            headers = [
                "번호",
                "시트명",
                "컬럼",
                "필드명",
                "규칙 내용",
                "조건",
                "비고",
                "AI 해석 여부",
                "AI 규칙 ID",
                "AI 규칙 유형",
                "AI 신뢰도"
            ]

            # Row 1: Main headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Row 2: Sub headers (empty or descriptive) - parser skips rows 1-2
            sub_headers = [
                "(자동)",
                "(시트명)",
                "(열)",
                "(필드)",
                "(검증 규칙)",
                "(조건)",
                "(비고)",
                "(AI)",
                "(규칙ID)",
                "(유형)",
                "(신뢰도)"
            ]
            sub_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            for col_idx, sub_header in enumerate(sub_headers, start=1):
                cell = ws.cell(row=2, column=col_idx, value=sub_header)
                cell.fill = sub_header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data rows starting from row 3 (matches parser min_row=3)
            for idx, rule in enumerate(all_rules, start=3):
                ws.cell(row=idx, column=1, value=idx - 2)  # 번호: 1, 2, 3...
                ws.cell(row=idx, column=2, value=rule.get('display_sheet_name'))
                ws.cell(row=idx, column=3, value=rule.get('column_letter'))
                ws.cell(row=idx, column=4, value=rule.get('field_name'))
                ws.cell(row=idx, column=5, value=rule.get('rule_text'))
                ws.cell(row=idx, column=6, value=rule.get('condition'))
                ws.cell(row=idx, column=7, value=rule.get('note'))
                ws.cell(row=idx, column=8, value="예" if rule.get('ai_rule_id') else "아니오")
                ws.cell(row=idx, column=9, value=rule.get('ai_rule_id') or "")
                ws.cell(row=idx, column=10, value=rule.get('ai_rule_type') or "")
                ws.cell(row=idx, column=11, value=rule.get('ai_confidence_score') or "")

            # Adjust column widths (11 columns now)
            column_widths = [8, 25, 10, 20, 40, 30, 30, 12, 20, 15, 10]
            for col_idx, width in enumerate(column_widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

            # Add metadata sheet
            ws_meta = wb.create_sheet("파일 정보")
            metadata_rows = [
                ["항목", "값"],
                ["파일명", file_record['file_name']],
                ["파일 버전", file_record.get('file_version', '')],
                ["업로드일시", str(file_record['uploaded_at'])],
                ["업로드자", file_record.get('uploaded_by', '')],
                ["총 규칙 수", file_record['total_rules_count']],
                ["시트 수", file_record['sheet_count']],
                ["상태", file_record['status']],
                ["비고", file_record.get('notes', '')]
            ]

            for row_idx, row_data in enumerate(metadata_rows, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws_meta.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment

            ws_meta.column_dimensions['A'].width = 20
            ws_meta.column_dimensions['B'].width = 50

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            excel_bytes = output.getvalue()
            print(f"[RuleService] Excel export completed ({len(excel_bytes)} bytes)")

            return excel_bytes

        except Exception as e:
            print(f"[RuleService] Error exporting to Excel: {str(e)}")
            raise Exception(f"Failed to export rules to Excel: {str(e)}")

    async def reinterpret_rules(
        self,
        file_id: str,
        use_local_parser: bool = True
    ) -> Dict[str, Any]:
        """
        저장된 원본 파일로 규칙 재해석

        Args:
            file_id: UUID string of the rule file
            use_local_parser: True면 로컬 파서 사용 (AI 오류 방지)

        Returns:
            Dict: 재해석 결과 통계
        """
        print(f"[RuleService] Starting re-interpretation for file: {file_id}")

        try:
            # Step 1: 원본 파일 조회
            original_content = await self.repository.get_original_file(UUID(file_id))

            if not original_content:
                raise Exception("원본 파일이 저장되어 있지 않습니다. 규칙 파일을 다시 업로드해주세요.")

            print(f"[RuleService] Loaded original file ({len(original_content)} bytes)")

            # Step 2: 원본 파일 파싱해서 자연어 규칙 추출
            print("[RuleService] Parsing original file...")
            natural_language_rules, sheet_row_counts, _, _ = parse_rules_from_excel(original_content)
            print(f"[RuleService] Parsed {len(natural_language_rules)} rules from original file")

            # Step 3: 로컬 파서로 규칙 해석
            print(f"[RuleService] Interpreting rules with local parser...")
            from ai_layer import AIRuleInterpreter
            interpreter = AIRuleInterpreter()
            validation_rules, conflicts = interpreter._local_rule_parser(natural_language_rules)
            print(f"[RuleService] Generated {len(validation_rules)} validation rules")

            # Step 4: 기존 규칙의 AI 해석 초기화
            print("[RuleService] Clearing existing AI interpretations...")
            cleared_count = await self.repository.clear_ai_interpretation(UUID(file_id))
            print(f"[RuleService] Cleared {cleared_count} rules")

            # Step 5: 새 해석 결과를 DB에 저장
            print("[RuleService] Saving new interpretations to DB...")
            db_rules = await self.repository.get_rules_by_file(UUID(file_id), active_only=True)

            updated_count = 0
            for db_rule in db_rules:
                # 필드명과 시트로 매칭
                field_name = db_rule.get('field_name')
                sheet_name = db_rule.get('canonical_sheet_name')

                # 매칭되는 해석 규칙 찾기
                for val_rule in validation_rules:
                    if val_rule.field_name == field_name and val_rule.source.sheet_name == sheet_name:
                        # AI 해석 데이터 업데이트
                        ai_data = {
                            "ai_rule_id": val_rule.rule_id,
                            "ai_rule_type": val_rule.rule_type,
                            "ai_parameters": val_rule.parameters,
                            "ai_error_message": val_rule.error_message_template,
                            "ai_interpretation_summary": val_rule.ai_interpretation_summary,
                            "ai_confidence_score": float(val_rule.confidence_score),
                            "ai_interpreted_at": datetime.now().isoformat(),
                            "ai_model_version": "local-parser"
                        }
                        success = await self.repository.update_rule_ai_interpretation(
                            UUID(db_rule['id']),
                            ai_data
                        )
                        if success:
                            updated_count += 1
                        break

            # Step 6: 해석 상태 업데이트
            engine = 'local'
            status = 'completed' if updated_count > 0 else 'failed'
            await self.repository.update_interpretation_status(UUID(file_id), status, engine)

            print(f"[RuleService] Re-interpretation completed: {updated_count}/{len(db_rules)} rules updated")
            return {
                'total_rules': len(db_rules),
                'interpreted_rules': updated_count,
                'skipped_rules': len(db_rules) - updated_count,
                'failed_rules': 0,
                'engine_used': engine,
                'original_file_available': True
            }

        except Exception as e:
            print(f"[RuleService] Error during re-interpretation: {str(e)}")
            await self.repository.update_interpretation_status(UUID(file_id), 'failed')
            raise Exception(f"Failed to re-interpret rules: {str(e)}")


# =============================================================================
# Testing
# =============================================================================

if __name__ == "__main__":
    import asyncio

    async def test_service():
        """Test service operations"""
        print("=" * 70)
        print("RuleService Test")
        print("=" * 70)

        try:
            service = RuleService()
            print("✓ Service initialized")

            # Test list files
            print("\nListing rule files...")
            files = await service.list_rule_files(limit=5)
            print(f"Found {len(files)} rule files")

            for file in files:
                print(f"  - {file.file_name} (ID: {file.id})")

            print("\n✓ Service test completed successfully")

        except Exception as e:
            print(f"\n✗ Service test failed: {str(e)}")

        print("=" * 70)

    # Run test
    asyncio.run(test_service())
