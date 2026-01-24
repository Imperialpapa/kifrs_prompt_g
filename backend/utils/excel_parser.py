"""
Excel Parsing Utilities
=======================
Excel 파일(규칙 파일)을 파싱하여 자연어 규칙 리스트를 추출하는 유틸리티 모듈
"""

import io
import re
from typing import List, Dict, Any, Tuple
from openpyxl import load_workbook


def normalize_sheet_name(name: str) -> str:
    """
    시트 이름 정규화
    - 줄바꿈, 탭 등을 공백으로 치환 (글자 붙음 방지)
    - 연속된 공백을 단일 공백으로 치환

    Args:
        name: 원본 시트 이름

    Returns:
        정규화된 시트 이름
    """
    if not isinstance(name, str):
        return str(name)

    # 제어 문자를 공백으로 치환 (빈 문자열이 아님!)
    normalized = name.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')

    # 연속된 공백 제거
    normalized = re.sub(r'\s+', ' ', normalized)

    return normalized.strip()


def get_canonical_name(name: str) -> str:
    """
    비교를 위한 정규화 (모든 공백 제거)

    Args:
        name: 원본 시트 이름

    Returns:
        Canonical 시트 이름 (공백 제거됨)
    """
    norm = normalize_sheet_name(name)
    return "".join(norm.split())


def get_visible_sheet_names(content: bytes) -> List[str]:
    """
    Excel 파일에서 숨겨지지 않은(Visible) 시트 이름 목록만 반환
    - .xlsx: openpyxl로 hidden 시트 제외
    - .xls: pandas로 모든 시트 반환 (hidden 여부 확인 불가)
    
    Args:
        content: Excel 파일의 바이트 내용
        
    Returns:
        List[str]: 숨겨지지 않은 시트 이름 목록 (또는 전체 목록)
    """
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        visible_sheets = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # sheet_state가 'visible'인 경우만 포함 (hidden, veryHidden 제외)
            if sheet.sheet_state == 'visible':
                visible_sheets.append(sheet_name)
        
        wb.close()
        return visible_sheets
    except Exception as e:
        # .xls 파일이거나 손상된 경우 등 openpyxl로 열 수 없을 때
        print(f"[ExcelParser] Warning: Could not check sheet visibility (likely .xls file). Falling back to all sheets. Error: {e}")
        try:
            import pandas as pd
            excel_file = pd.ExcelFile(io.BytesIO(content))
            return excel_file.sheet_names
        except Exception as pd_e:
            print(f"[ExcelParser] Error: Failed to list sheets with pandas: {pd_e}")
            return []


def parse_rules_from_excel(content: bytes) -> Tuple[List[Dict[str, Any]], Dict[str, int], int, int]:
    """
    Excel B 파일(규칙 파일)을 파싱하여 자연어 규칙 리스트를 반환

    Args:
        content: Excel 파일의 바이트 내용

    Returns:
        Tuple containing:
        - natural_language_rules: 파싱된 규칙 리스트
        - sheet_row_counts: 시트별 행 개수 (display name 기준)
        - total_raw_rows: 전체 원본 행 수
        - reported_max_row: 엑셀 파일이 메타데이터로 보고하는 총 행 수 (헤더 제외)
    """
    wb = load_workbook(io.BytesIO(content), data_only=True)
    natural_language_rules = []
    sheet_row_counts = {}
    total_raw_rows = 0
    reported_max_row = 0

    # 메타데이터 시트 제외 목록 (규칙 매핑에서 제외)
    EXCLUDED_SHEETS = {'파일 정보', '파일정보', 'File Info', 'Metadata', 'metadata', '_metadata'}

    print(f"   [INFO] Found {len(wb.sheetnames)} sheets in rules file: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        # 메타데이터 시트는 건너뛰기
        if sheet_name in EXCLUDED_SHEETS or sheet_name.startswith('_'):
            print(f"   [INFO] Skipping metadata sheet: '{sheet_name}'")
            continue
        ws = wb[sheet_name]
        print(f"   [INFO] Processing rules sheet: '{sheet_name}' (Reported Max Row: {ws.max_row})")

        # 메타데이터 상의 max_row 누적 (헤더 2행 제외)
        if ws.max_row > 2:
            reported_max_row += (ws.max_row - 2)

        last_sheet_name_val = None
        consecutive_empty_rows = 0

        for row_idx, row_values in enumerate(ws.iter_rows(min_row=3, max_row=1000, values_only=True), start=3):
            if all(cell is None for cell in row_values):
                consecutive_empty_rows += 1
                if consecutive_empty_rows >= 5:
                    break
                continue

            consecutive_empty_rows = 0
            total_raw_rows += 1

            raw_sheet_val = row_values[1] if len(row_values) > 1 else None
            if isinstance(raw_sheet_val, str) and not raw_sheet_val.strip():
                raw_sheet_val = None
            if raw_sheet_val is not None:
                last_sheet_name_val = raw_sheet_val

            # Fallback to actual worksheet name if column B is empty
            # This handles cases where rules are separated by tabs but column B is left blank
            current_sheet_name = last_sheet_name_val if last_sheet_name_val else sheet_name

            if current_sheet_name:
                disp_name = normalize_sheet_name(str(current_sheet_name))
                sheet_row_counts[disp_name] = sheet_row_counts.get(disp_name, 0) + 1

            if not current_sheet_name:
                continue

            field_name = row_values[3] if len(row_values) > 3 else None
            condition = row_values[5] if len(row_values) > 5 else None
            if condition and "해당없음" in str(condition):
                continue

            column_letter = row_values[2] if len(row_values) > 2 else ""
            validation_rule = row_values[4] if len(row_values) > 4 else ""
            note = row_values[6] if len(row_values) > 6 else ""
            safe_field_name = str(field_name) if field_name else "(필드명 없음)"
            rule_text = str(validation_rule) if validation_rule else (f"조건: {condition}" if condition else f"기본 검증 ({safe_field_name})")

            rule_entry = {
                "sheet": get_canonical_name(str(current_sheet_name)),
                "display_sheet_name": normalize_sheet_name(str(current_sheet_name)),
                "row": row_idx,
                "column_letter": str(column_letter) if column_letter else "",
                "field": safe_field_name,
                "rule_text": rule_text,
                "condition": str(condition) if condition else "",
                "note": str(note) if note else ""
            }

            # CRITICAL DEBUG: Print first few rules to verify field-rule mapping
            if len(natural_language_rules) < 10:
                print(f"   [DEBUG] Rule #{len(natural_language_rules)+1}: field='{safe_field_name}' rule_text='{rule_text[:50]}'")

            natural_language_rules.append(rule_entry)

    return natural_language_rules, sheet_row_counts, total_raw_rows, reported_max_row
